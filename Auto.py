from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep, strftime
from random import randint
import pandas as pd
from openpyxl import load_workbook
import os

dirname = 'D:\\Report\\1398\\TabdilVaz'
progdirname = os.path.dirname(__file__)
       
chromedriver_path = r'C://Users/a.akhavan\AppData/Roaming/npm/node_modules/chromedriver/lib/chromedriver/chromedriver.exe' # Change this to your own chromedriver path!
webdriver = webdriver.Chrome(executable_path=chromedriver_path)

webdriver.get('http://amarnameh.imo.org.ir')

username = webdriver.find_element_by_name('txtUsername')
username.send_keys('')
password = webdriver.find_element_by_name('txtPassword')
password.send_keys('')

webdriver.find_element_by_xpath('//*[@id="ctl01"]/div[6]/div').click()

companies = ['SH']

# companies = ['SazMotori','SazMotori']
forms = ['Frm2','Frm3']
i = 0
for company in companies:
    wb = load_workbook(os.path.join(dirname, company+'\\Frm.xlsx'))

    if('Frm2' in forms):
        ws = wb['Frm2']

        added = 0
        withError = 0
        col = []
        values = []

        ws['AZ1'] = 'state'
        ws['BA1'] = 'error'
        ws['BB1'] = 'place'
        ws['BC1'] = 'serial'
        # http://amarnameh.imo.org.ir/Input/Update.aspx?Id=8023&cid=281
        index = 0
        total = ws.max_row
        for rownum in ws.iter_rows():
            index = index + 1
            try:
                if index == 1:
                    col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                else:
                    values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]

                    if(values[51] == 'added'):
                        continue
                    elif(values[54] != ''):
                        webdriver.get('http://amarnameh.imo.org.ir/Input/Update.aspx?Id=8023&cid='+values[54])
                    else:
                        webdriver.get('http://amarnameh.imo.org.ir/Input/EditEx.aspx?Id=8023')

                    webdriver.find_element_by_name('ctl00_ContentPlaceHolder1_dialog_233104304').click()         
                    webdriver.find_element_by_id('ctl00_ContentPlaceHolder1_FACT_FIELD_44788t3').click()#isfahan
                    i = 6
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44789"]/option['+values[i]+']').click()# وضعیت استخدامی
                    i = 7
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44790"]/option['+values[i]+']').click()# مامور
                    i = 8
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44791"]/option['+values[i]+']').click()# از
                    i = 9
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44792"]/option['+values[i]+']').click()# نوع مامور
                    i = 10
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44793').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44793').send_keys(values[i])#نام
                    i = 11
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44794').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44794').send_keys(values[i])#نام خانوادگی
                    i = 12
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44795').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44795').send_keys(values[i])#نام پدر
                    i = 13
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44796').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44796').send_keys(values[i])#کدملی
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 2

                    i = 14
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44797"]/option['+values[i]+']').click()# آخرين مدرک تحصیلی
                    i = 15
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44798"]/option['+values[i]+']').click()# رشته تحصيلي
                    i = 16
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44799').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44799').send_keys(values[i])# ساير رشته تحصيلي
                    i = 17
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44800"]/option['+values[i]+']').click()# جنسیت
                    i = 18
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44801').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44801').send_keys(values[i])# تاريخ تولد
                    i = 19
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44802"]/option['+values[i]+']').click()# وضعيت تاهل
                    i = 20
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44803').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44803').send_keys(values[i])# تعداد فرزند
                    i = 21
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44804').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44804').send_keys(values[i])# شماره تلفن همراه
                    i = 22
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44805').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44805').send_keys(values[i])# شماره تلفن محل كار همراه با كد تلفن شهري
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 3

                    i = 23
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44806"]/option['+values[i]+']').click()# سابقه ايثارگري ؟
                    if(values[i]=='2'):
                        i = 24
                        webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44807"]/option['+values[i]+']').click()# وضعيت ايثارگري
                        i = 25
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44808').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44808').send_keys(values[i])# تاريخ شروع رزمندگي
                        i = 26
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44809').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44809').send_keys(values[i])# تاريخ پايان رزمندگيi = 2
                        i = 27
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44810').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44810').send_keys(values[i])# تاريخ جانبازي
                        i = 28
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44811').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44811').send_keys(values[i])# درصد جانبازي
                        i = 29
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44812').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44812').send_keys(values[i])# تاریخ شروع اسارت
                        i = 30
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44813').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44813').send_keys(values[i])# تاریخ پایان اسارت
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 4

                    i = 31
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44814"]/option['+values[i]+']').click()# نوع صندوق
                    i = 32
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44815"]/option['+values[i]+']').click()# بيمه
                    i = 33
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44816').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44816').send_keys(values[i])# كد بيمه
                    i = 34
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44817').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44817').send_keys(values[i])# نحوه بكارگيري نيرو
                    i = 35
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44818').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44818').send_keys(values[i])# تاريخ ورود به شهرداري
                    i = 36
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44819').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44819').send_keys(values[i])# شماره مستخدم در صورت دارا بودن
                    i = 37
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44820"]/option['+values[i]+']').click()# وضعیت اشتغال
                    i = 38
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44821"]/option['+values[i]+']').click()# تعهد استخدامي(فقط براي پرسنل پيماني با آزمون)
                    i = 39
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44822"]/option['+values[i]+']').click()# پست سازماني
                    if(values[i] == '2'):
                        i = 40
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44823').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44823').send_keys(values[i])# عنوان پست سازمانی
                        i = 41
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44824').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44824').send_keys(values[i])# شماره پست سازمانی
                    i = 42
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44825').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44825').send_keys(values[i])# عنوان شغل سازماني
                    i = 43
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44826').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44826').send_keys(values[i])# رشته شغلي
                    i = 44
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44827').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44827').send_keys(values[i])# رسته شغلي
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 5

                    i = 45
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44828').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44828').send_keys(values[i])# گروه
                    i = 46
                    strGoroh = '1'
                    if(values[i] != ''):
                        strGoroh = str(int(values[i])+2)
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44829"]/option['+strGoroh+']').click()# گروه تشويقي
                    i = 47
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44830"]/option['+values[i]+']').click()# محل خدمت
                    if(values[47]=='6'):
                        i = 48
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44831').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44831').send_keys(values[i])# عنوان ساير محل خدمت
                    if(values[47]=='4'):
                        i = 49
                        webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44832"]/option['+values[i]+']').click()# نام سازمان وابسته به شهرداري
                        i = 50
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44833').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44833').send_keys(values[i])# نام ساير سازمان وابسته به شهرداري
                    
                    temp = os.path.join(progdirname,'1234567890.pdf')
                    attachmentError = ''
                    i=-1
                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'t.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' t'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44834').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44834').send_keys(attachmentFile)# مدرک تحصیلی

                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'s.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' s'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44835').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44835').send_keys(attachmentFile)# سابقه بیمه

                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'g.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' g'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44836').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44836').send_keys(attachmentFile)# قزازداد

                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'m.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' m'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44837').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44837').send_keys(attachmentFile)# مجوز به کارگیری

                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_mnuSave"]').click()# finish
                    added += 1
                    if(attachmentError == ''):
                        ws['AZ'+str(index)] = "added"
                        print('{0} - {1} {2:3.0f}% Frm2 Adeed {3} '.format(index,total,index/total*100,values[13]))
                    else:
                        ws['AZ'+str(index)] = "added"
                        ws['BA'+str(index)] = 'atachment'
                        ws['BB'+str(index)] = attachmentError
                        print('{0} - {1} {2:3.0f}% Frm2 Adeed AttachmentError {3} - {4}'.format(index,total,index/total*100,values[13], attachmentError))
                    try:
                        webdriver.switch_to.alert.accept()
                    except:
                        ws['AZ'+str(index)] = "error"
            
            except:
                ws['AZ'+str(index)] = "error"
                if i > 0 :
                    ws['BA'+str(index)] = col[i]
                    ws['BB'+str(index)] = values[i]
                    print('{0} - {1} {2:3.0f}% Frm2 Error {3} - {4} {5} {6} '.format(index,total,index/total*100,values[13], i,col[i], values[i]))                    
                else : 
                    ws['BA'+str(index)] = 'atachment'
                    ws['BB'+str(index)] = i
                    print('{0} - {1} {2:3.0f}% Frm2 Error Frm2 Error AttachmentError {3} {4}'.format(index,total,index/total*100,values[13], attachmentError))
                
                withError += 1
                continue

        wb.save(os.path.join(dirname, company+'\\Frm.xlsx'))

        print('Frm2 {} Added {} Person.'.format(company,added))
        print('Frm2 {} With {} Error.'.format(company,withError))

    if('Frm3' in forms):
        ws = wb['Frm3']
        added = 0
        withError = 0
        col = []
        values = []

        ws['AJ1'] = 'state'
        ws['AK1'] = 'error'
        ws['AL1'] = 'place'
        ws['AM1'] = 'serial'
        # http://amarnameh.imo.org.ir/Input/Update.aspx?Id=8023&cid=281
        index = 0
        total = ws.max_row
        for rownum in ws.iter_rows():
            index = index + 1
            try:
                if index == 1:
                    col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                else:
                    values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]

                    if(values[36] == 'added'):
                        continue
                    elif(values[39] != ''):
                        webdriver.get('http://amarnameh.imo.org.ir/Input/EditEx.aspx?Id=8024&cid='+values[54])
                    else:
                        webdriver.get('http://amarnameh.imo.org.ir/Input/EditEx.aspx?Id=8024')

                    webdriver.find_element_by_name('ctl00_ContentPlaceHolder1_dialog_954246153').click()         
                    webdriver.find_element_by_id('ctl00_ContentPlaceHolder1_FACT_FIELD_44838t3').click()#isfahan
                    i = 6
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44839').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44839').send_keys(values[i])#نام
                    i = 7
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44840').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44840').send_keys(values[i])#نام خانوادگی
                    i = 8
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44841').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44841').send_keys(values[i])#نام پدر
                    i = 9
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44842').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44842').send_keys(values[i])#کدملی
                    i = 10
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44843"]/option['+values[i]+']').click()# آخرين مدرک تحصیلی
                    i = 11
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44844"]/option['+values[i]+']').click()# رشته تحصيلي
                    i = 12
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44845').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44845').send_keys(values[i])# ساير رشته تحصيلي
                    i = 13
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44846"]/option['+values[i]+']').click()# جنسیت
                    i = 14
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44847').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44847').send_keys(values[i])# تاريخ تولد
                    i = 15
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44848"]/option['+values[i]+']').click()# وضعيت تاهل
                    i = 16
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44849').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44849').send_keys(values[i])# تعداد فرزند
                    i = 17
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44850').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44850').send_keys(values[i])# شماره تلفن همراه
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 2

                    i = 18
                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44851"]/option['+values[i]+']').click()# سابقه ايثارگري ؟
                    if(values[i]=='2'):
                        i = 19
                        webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_FACT_FIELD_44852"]/option['+values[i]+']').click()# وضعيت ايثارگري
                        i = 20
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44853').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44853').send_keys(values[i])# تاريخ شروع رزمندگي
                        i = 21
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44854').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44854').send_keys(values[i])# تاريخ پايان رزمندگيi = 2
                        i = 22
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44855').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44855').send_keys(values[i])# تاريخ جانبازي
                        i = 23
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44856').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44856').send_keys(values[i])# درصد جانبازي
                        i = 24
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44857').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44857').send_keys(values[i])# تاریخ شروع اسارت
                        i = 25
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44858').clear()
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44858').send_keys(values[i])# تاریخ پایان اسارت
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 3

                    i = 26
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44859').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44859').send_keys(values[i])# سوابق شركتي در شهرداري از تاريخ
                    i = 27
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44860').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44860').send_keys(values[i])# سوابق شركتي در شهرداري تا تاريخ(29-4-98
                    i = 28
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44861').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44861').send_keys(values[i])# مجموع سوابق شركتي در شهرداري -سال
                    i = 29
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44862').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44862').send_keys(values[i])# مجموع سوابق شركتي در شهرداري -ماه
                    i = 30
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44863').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44863').send_keys(values[i])# مجموع سوابق شركتي در شهرداري -روز
                    i = 31
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44864').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44864').send_keys(values[i])# محل دقيق خدمت
                    i = 32
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44865').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44865').send_keys(values[i])# شغل فعلي در شهرداري
                    i = 33
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44866').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44866').send_keys(values[i])# نام شركت پيمانكاري مربوطه
                    i = 34
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44867').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44867').send_keys(values[i])# كد بيمه تامين اجتماعي
                    webdriver.find_element_by_xpath('//*[@id="mnuNext"]').click()# next page 4

                    temp = os.path.join(progdirname,'1234567890.pdf')
                    attachmentError = ''
                    i=-1
                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'t.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' t'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44868').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44868').send_keys(attachmentFile)# مدرک تحصیلی

                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'s.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' s'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44869').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44869').send_keys(attachmentFile)# سابقه بیمه

                    attachmentFile = os.path.join(dirname,company+'\\'+values[13]+'g.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' g'
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44870').clear()
                    webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$FACT_FIELD_44870').send_keys(attachmentFile)# قزازداد        

                    webdriver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_mnuSave"]').click()# finish
                    added += 1
                    if(attachmentError == ''):
                        ws['AJ'+str(index)] = "added"
                        print('{0} - {1} {2:3.0f}% Frm 3 Adeed {3} '.format(index,total,index/total*100,values[13]))
                    else:
                        ws['AJ'+str(index)] = "added"
                        ws['AK'+str(index)] = 'atachment'
                        ws['AL'+str(index)] = attachmentError
                        print('{0} - {1} {2:3.0f}% Frm3 Adeed AttachmentError {3} - {4}'.format(index,total,index/total*100,values[13], attachmentError))
                    try:
                        webdriver.switch_to.alert.accept()
                    except:
                        ws['AJ'+str(index)] = "error"
            
            except:
                ws['AJ'+str(index)] = "error"
                if i > 0 :
                    ws['AK'+str(index)] = col[i]
                    ws['AL'+str(index)] = values[i]
                    print('{0} - {1} {2:3.0f}% Error {3} - {4} {5} {6}'.format(index,total,index/total*100,values[13], i,col[i], values[i]))                    
                else : 
                    ws['AK'+str(index)] = 'atachment'
                    ws['AL'+str(index)] = i
                    print('{0} - {1} {2:3.0f}% Error AttachmentError {3} {4}'.format(index,total,index/total*100,values[13],attachmentError))
                
                withError += 1
                continue

        wb.save(os.path.join(dirname, company+'\\Frm.xlsx'))

        print('Frm3 {} Added {} Person.'.format(company,added))
        print('Frm3 {} With {} Error.'.format(company,withError))