from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep, strftime
from random import randint
import pandas as pd
from openpyxl import load_workbook
from time import sleep
import os

dirname = 'D:\\Report\\1398'
progdirname = os.path.dirname(__file__)
       
chromedriver_path = r'C://Users/a.akhavan\AppData/Roaming/npm/node_modules/chromedriver/lib/chromedriver/chromedriver.exe' # Change this to your own chromedriver path!
webdriver = webdriver.Chrome(executable_path=chromedriver_path)

webdriver.get('http://amarnameh.imo.org.ir')

username = webdriver.find_element_by_name('txtUsername')
username.send_keys('')
password = webdriver.find_element_by_name('txtPassword')
password.send_keys('')

webdriver.find_element_by_xpath('//*[@id="ctl01"]/div[6]/div').click()

companies = ['Kol']

# companies = ['SazMotori','SazMotori']
forms = ['Frm3']
i = 0
attachmentError = ''
for company in companies:
    wb = load_workbook(os.path.join(dirname,'FrmT - Duplicate.xlsx'))

    if('Frm3' in forms):
        ws = wb['Frm3']
        added = 0
        withError = 0
        col = []
        values = []

        index = 0
        total = ws.max_row
        for rownum in ws.iter_rows():
            index = index + 1
            try:
                if index == 1:
                    col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                else:
                    values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                    webdriver.get('http://amarnameh.imo.org.ir/Input/Records.aspx?Id=8024&ShowFilter=true')

                    try:
                        while webdriver.find_element_by_id('ctl00_ContentPlaceHolder1_grdReadOnly_ctl00_ctl06_number_Label_lab').get_attribute('innerHTML') == '2':
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_id').clear()                     
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_id').send_keys(values[3]) 
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_id').clear()                     
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_id').send_keys(values[3]) 
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_FACT_FIELD_44838_0_VALUE').clear()
                            webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl02$ctl03$FilterTextBox_FACT_FIELD_44838_0_VALUE').send_keys('اصفهان')                     
                            sleep(7) 
                    except:                          
                        webdriver.find_element_by_name('ctl00$ContentPlaceHolder1$grdReadOnly$ctl00$ctl04$gbccolumn1').click()                    
                        webdriver.switch_to.alert.accept()
                        webdriver.switch_to.alert.accept()
                        print('{0} - {1} {2:3.0f}% '.format(index,total,index/total*100))
                        webdriver.get('http://amarnameh.imo.org.ir/Input/Records.aspx?Id=8024&ShowFilter=true')
    
            except:
                print('{0} - {1} {2:3.0f}% ERROR'.format(index,total,index/total*100))                
                withError += 1
                continue

        print('Frm3 {} Added {} Person.'.format(company,added))
        print('Frm3 {} With {} Error.'.format(company,withError))